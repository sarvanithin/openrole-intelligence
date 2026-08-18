"""Strict importer for official DOL H-1B LCA disclosure data."""

from __future__ import annotations

import csv
import hashlib
import re
from collections import defaultdict
from collections.abc import Iterable, Mapping
from datetime import UTC, datetime
from itertools import chain
from pathlib import Path

from fortune_intel.storage import JobRepository
from fortune_intel.storage.identity import normalize_company_name

DOL_SOURCE_URL = "https://www.dol.gov/agencies/eta/foreign-labor/performance"

_COLUMNS = {
    "case": ("CASE_NUMBER", "Case Number"),
    "employer": ("EMPLOYER_NAME", "Employer Name"),
    "status": ("CASE_STATUS", "Case Status"),
    "visa": ("VISA_CLASS", "Visa Class"),
    "workers": ("TOTAL_WORKER_POSITIONS", "TOTAL_WORKERS", "Total Worker Positions"),
}


def _column_map(row: Mapping[str, object]) -> dict[str, str]:
    actual = {str(key).casefold().strip(): str(key) for key in row}
    resolved = {}
    for logical, aliases in _COLUMNS.items():
        for alias in aliases:
            if alias.casefold() in actual:
                resolved[logical] = actual[alias.casefold()]
                break
    missing = sorted(set(_COLUMNS) - set(resolved))
    if missing:
        raise ValueError(f"DOL disclosure is missing required columns: {', '.join(missing)}")
    return resolved


def _text(row: Mapping[str, object], column: str) -> str:
    value = row.get(column)
    if value is None or str(value).strip().casefold() in {"", "nan", "none"}:
        return ""
    return str(value).strip()


def _rows(path: Path) -> Iterable[Mapping[str, object]]:
    if path.suffix.casefold() == ".csv":
        with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
            yield from csv.DictReader(handle)
        return
    if path.suffix.casefold() != ".xlsx":
        raise ValueError("DOL disclosure file must be CSV or XLSX")
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        values = workbook.active.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values, ())]
        for row in values:
            yield dict(zip(headers, row, strict=False))
    finally:
        workbook.close()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _validate_fiscal_year(path: Path, fiscal_year: int) -> None:
    if not 2000 <= fiscal_year <= datetime.now(UTC).year + 1:
        raise ValueError("fiscal_year is outside the supported range")
    match = re.search(r"FY[_ -]?(20\d{2})", path.name, re.IGNORECASE)
    if not match:
        raise ValueError("DOL filename must include its fiscal year, for example FY2025")
    if int(match.group(1)) != fiscal_year:
        raise ValueError("--fiscal-year conflicts with the fiscal year in the filename")


def import_dol_lca(
    repository: JobRepository,
    disclosure_path: str | Path,
    *,
    fiscal_year: int,
) -> dict[str, int]:
    """Import certified H-1B worker positions as provisional employer matches."""

    path = Path(disclosure_path)
    _validate_fiscal_year(path, fiscal_year)
    checksum = _sha256_file(path)
    iterator = iter(_rows(path))
    first = next(iterator, None)
    if first is None:
        raise ValueError("DOL disclosure file is empty")
    columns = _column_map(first)

    aggregate: dict[str, dict[str, object]] = {}
    seen_cases: set[str] = set()
    stats = defaultdict(int)
    for row in chain((first,), iterator):
        stats["rows_read"] += 1
        case_number = _text(row, columns["case"])
        if not case_number or case_number in seen_cases:
            stats["rows_skipped_duplicate_or_missing_case"] += 1
            continue
        seen_cases.add(case_number)
        if _text(row, columns["visa"]).upper() != "H-1B":
            stats["rows_skipped_other_visa_class"] += 1
            continue
        status = _text(row, columns["status"]).upper()
        if status == "CERTIFIED-WITHDRAWN":
            stats["rows_skipped_certified_withdrawn"] += 1
            continue
        if status != "CERTIFIED":
            stats["rows_skipped_non_certified"] += 1
            continue
        employer = _text(row, columns["employer"])
        workers_text = _text(row, columns["workers"]).replace(",", "")
        try:
            workers_float = float(workers_text)
            workers = int(workers_float)
            if workers_float != workers or workers < 0:
                raise ValueError
        except ValueError:
            stats["rows_skipped_invalid_worker_positions"] += 1
            continue
        if not employer:
            stats["rows_skipped_missing_employer"] += 1
            continue
        normalized_employer = normalize_company_name(employer)
        entry = aggregate.setdefault(
            normalized_employer, {"employer_name": employer, "worker_positions": 0}
        )
        entry["worker_positions"] = int(entry["worker_positions"]) + workers

    imported_at = datetime.now(UTC).replace(microsecond=0).isoformat()
    for entry in aggregate.values():
        repository.upsert_h1b_employer(
            str(entry["employer_name"]),
            fiscal_year=fiscal_year,
            lca_worker_positions=int(entry["worker_positions"]),
            source="dol_lca",
            source_url=DOL_SOURCE_URL,
            source_document=path.name,
            source_checksum=checksum,
            imported_at=imported_at,
        )

    matched_by_company: dict[int, int] = defaultdict(int)
    for entry in aggregate.values():
        employer_name = str(entry["employer_name"])
        worker_positions = int(entry["worker_positions"])
        company = repository.find_company_by_normalized_name(employer_name)
        if company is None:
            stats["employers_unmatched"] += 1
            continue
        matched_by_company[int(company["id"])] += worker_positions

    for company_id, worker_positions in matched_by_company.items():
        repository.record_sponsorship_fact(
            company_id,
            source="dol_lca",
            fiscal_year=fiscal_year,
            lca_worker_positions=worker_positions,
            # Normalized display-name equality is only a candidate link. A
            # reviewed alias/address/domain match is required to reach 0.90.
            entity_match_confidence=0.85,
            source_url=DOL_SOURCE_URL,
            source_document=path.name,
            source_checksum=checksum,
            match_method="provisional_normalized_name",
        )
    stats["employers_aggregated"] = len(aggregate)
    stats["h1b_employers_indexed"] = len(aggregate)
    stats["companies_provisionally_matched"] = len(matched_by_company)
    return dict(stats)
